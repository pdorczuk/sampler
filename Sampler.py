import sys
from tkinter import Tk, messagebox
from tkinter.filedialog import askopenfilename
from tkinter import *

import os
import re
from collections import defaultdict
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Color, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import random
import itertools

# Prod runs as an exe so the filepaths need to be set differently. Test uses the script directory.
if getattr(sys, 'frozen', False):  # We are running from an exe
    Tk().withdraw()  # Don't need a full GUI, so keep the root window from appearing
    # Open a file-browser to select the input file and return the path
    inventory_file = askopenfilename()
    base_path = sys._MEIPASS
    # this file is embedded in the EXE file
    database_file = os.path.join(base_path, 'requests.xlsx')
else:
    inventory_file = 'example_input.xlsx'  # testing input file
    database_file = 'requests.xlsx'


def main():
    inventory, client_name, audit_type = read_inventory(inventory_file)
    request_data, irl_references = create_request_list(inventory, audit_type)
    write_request_list(request_data, client_name)
    write_tracking_spreadsheet_hosts(
        request_data, irl_references, client_name, inventory, audit_type)


def read_inventory(inventory_file):
    """ Read a spreadsheet containing a list of servers and convert to a list
    :param inventory_file: Spreadsheet containing the server list
    :return inventory: Dictionary with platform as the key and hostnames as values
    """
    inventory = defaultdict(list)
    wb = openpyxl.load_workbook(inventory_file, read_only=True, data_only=True)
    sheet = wb['Sample']

    # Get the file processing info, the script cannot operate without an audit type so fail if its not a supported value
    client_name = sheet['B4'].value
    audit_type = sheet['B5'].value
    supported_audit_types = ['PCI', 'ISO', 'SOC', 'HIPAA', 'HITRUST']
    if audit_type not in map(str.upper, supported_audit_types):
        wb, sheet = open_report_workbook()
        sheet.cell(
            row=1, column=1).value = 'Unsupported audit type entered. Use a valid type as noted in the input template.'
        wb.save(client_name + '-Requests.xlsx')

    # Read the inventory platforms and hostnames
    for row in sheet.iter_rows(min_col=1, min_row=9, max_col=2, max_row=sheet.max_row):
        if row[0].value is not None:
            # Slice cell A from row tuple to form the key
            if row[:1][0].value in inventory.keys():
                # Slice cell B from row tuple to form the value
                inventory[row[:1][0].value] += [row[1:2][0].value]
            else:
                inventory[row[:1][0].value] = [row[1:2][0].value]
    return inventory, client_name, audit_type


def create_request_list(inventory, audit_type):
    """ Read the request spreadsheet and generate the request list. 
    :param inventory: Dictionary containing the sampled system inventory
    :return: 
    """
    request_data = {}
    irl_references = defaultdict(list)
    wb = openpyxl.load_workbook(database_file, read_only=True, data_only=True)

    counter = 1
    for k, v in inventory.items():
        ''' The requests database has a tab per platform with all the requests in it. so this section iterates through each dictionary key and checks if there is a corresponding tab in the spreadsheet. If there is it grabs the info and if there is no match then the generic tab is used as a catchall. '''
        sheet = ''
        if k in wb.sheetnames:
            sheet = wb[k]
        else:
            sheet = wb['generic']

        audit_columns = set()
        for row in sheet.iter_rows(min_col=1, min_row=1, max_col=sheet.max_column, max_row=sheet.max_row):
            request_data[counter] = {}

            for cell in row:
                if (cell.value is not None) and (row[0].value is not None):
                    if cell.row == 1:
                        # get all the global columns and any in row 1 that match the audit_type
                        if 'Global' in cell.value or audit_type in cell.value:
                            audit_columns.add(cell.column)
                    else:
                        # TODO create nested dict with ref as dict name and all other columns as values with the header as the key
                        if cell.column in audit_columns:
                            request_data[counter][(sheet.cell(1, cell.column).value)] = str(
                                cell.value).replace('%%', ", ".join(v))
                        if 'Reference' in sheet.cell(1, cell.column).value:
                            for item in v:
                                irl_references[item].append(cell.value)
            counter += 1
    return request_data, irl_references


def write_request_list(request_data, client_name):
    """ Writes the excel file with values to copy into AuditSource. """
    column_headers = ['Title', 'Instructions', 'Due Date', 'Reference', 'Population Request', 'Evidence Form',
                      'Related Section(s)', 'Assignee 1', 'Assignee 2', 'Assignee 3', 'Assignee 4', 'Assignee 5', 
                      'Project 1', 'Project 2', 'Project 3', 'Project 4', 'Project 5']
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Request List'
    # write the title row
    for col, val in enumerate(column_headers, start=1):
        sheet.cell(row=1, column=col).value = val
    # Write the table data into the spreadsheet
    row_num = 2
    for d_id, d_info in request_data.items():
        for row in sheet.iter_rows(min_col=1, min_row=row_num, max_col=len(column_headers), max_row=row_num):
            for key in d_info:
                for cell in row:
                    if (sheet.cell(1, cell.column).value).lower() in key.lower():
                        cell.value = d_info[key]
                    elif 'Population Request' in sheet.cell(1, cell.column).value:
                        cell.value = 'No'
                    elif 'Evidence Form' in sheet.cell(1, cell.column).value:
                        cell.value = 'Documentation'
                    else:
                        continue
            if row[0].value is not None:
                row_num += 1
    wb.save(client_name + ' - AuditSource Request List.xlsx')


def write_tracking_spreadsheet_hosts(request_data, irl_references, client_name, inventory, audit_type):
    """ use the info gathered so far to create an auditor tracking spreadsheet """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Tracking'
    column_num = 2
    row_num = 2
    longest = 0
    # Calculate the appropriate column width based on the longest hostname
    for item in inventory.values():
        for host in item:
            if len(host) > longest:
                longest = len(host) + 3

    # Writes all the IRL references down the first column
    for item in sorted(set(re.findall(r'[A-Z]+\-[0-9]{3}', str(request_data)))):
        sheet.cell(row=row_num, column=1).value = item
        row_num += 1

    def format_hostname_row(column_num, host, color, longest):
        sheet.cell(row=1, column=column_num).value = host
        sheet.cell(row=1, column=column_num).font = make_bold
        sheet.cell(row=1, column=column_num).fill = color
        column_letter = get_column_letter(column_num)
        sheet.column_dimensions[column_letter].width = longest
        column_num += 1
        return column_num

    # make the header row taller
    sheet.row_dimensions[1].height = 30

    # write the hostnames across the top row
    for k, v in inventory.items():
        color = random.choice(all_colors)
        all_colors.remove(color)
        for hostname in v:
            column_num = format_hostname_row(
                column_num, hostname, color, longest)

    ''' get list of all irl refs down the first column and iterate through the dictionaries in request_data
    and make a list of all the applicable IRL refs in each. then write N/A blocks on anything in the first column thats not
    on the list. basically, if its not applicable, then assume its not. '''
    hit_count = 0
    for k, v in irl_references.items():
        for row in sheet.iter_rows(min_col=2, min_row=2, max_col=sheet.max_column, max_row=sheet.max_row):
            for cell in row:
                try:
                    # if the hostname is in the top row of the column
                    if (sheet.cell(1, cell.column).value).lower() == k.lower():
                        # if the IRL reference is NOT in the list of values for that host
                        if sheet.cell(cell.row, 1).value not in re.findall(r'[A-Z]+\-[0-9]{3}', str(v)):
                            cell.value = 'N/A'
                            cell.fill = light_gray_fill
                except AttributeError:
                    pass

    wb.save(client_name + ' - System Evidence Tracking.xlsx')


# Constants
make_bold = Font(bold=True)
make_italic = Font(italic=True)
dark_green_fill = PatternFill(
    start_color='03990f', end_color='03990f', fill_type='solid')
light_gray_fill = PatternFill(
    start_color='bcb7b9', end_color='bcb7b9', fill_type='solid')
light_purple_fill = PatternFill(
    start_color='d2a2f2', end_color='d2a2f2', fill_type='solid')
light_green_fill = PatternFill(
    start_color='8af202', end_color='8af202', fill_type='solid')
light_pink_fill = PatternFill(
    start_color='f7b2f2', end_color='f7b2f2', fill_type='solid')
light_blue_fill = PatternFill(
    start_color='02f2ee', end_color='02f2ee', fill_type='solid')
light_orange_fill = PatternFill(
    start_color='efc25f', end_color='efc25f', fill_type='solid')
light_yellow_fill = PatternFill(
    start_color='eff24d', end_color='eff24d', fill_type='solid')
lime_green_fill = PatternFill(
    start_color='ccffcc', end_color='ccffcc', fill_type='solid')
bright_pink_fill = PatternFill(
    start_color='ffccff', end_color='ffccff', fill_type='solid')
all_colors = []
for item in dark_green_fill, light_purple_fill, light_green_fill, light_pink_fill, light_blue_fill, light_orange_fill, light_yellow_fill, lime_green_fill, bright_pink_fill:
    all_colors.append(item)
no_fill = openpyxl.styles.PatternFill(fill_type=None)

if __name__ == '__main__':
    main()

# Utility Functions
#######################################################################################################################
def open_report_workbook():
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = 'Request List'
    return wb, sheet
